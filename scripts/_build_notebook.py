"""Programmatically generate PoliMillionaire.ipynb.

Run from repo root:
    python scripts/_build_notebook.py

Produces a clean, sectioned notebook designed as an experimentation
workbench. Sections:
  0. Setup
  1. Configure
  2. Run
  3. Compare
  4. Save

Each section: markdown intro → code cell(s) → observations placeholder.
Code cells use Yoda-style comments (assignment style requirement);
markdown is plain English so the analytical narrative reads naturally.

This builder is internal — the deliverable is the .ipynb, not this file.
"""
from __future__ import annotations

import textwrap
from pathlib import Path

import nbformat as nbf


# ──────────────────────────────────────────────────────────────────────
# Cell helpers
# ──────────────────────────────────────────────────────────────────────

def md(text: str):
    return nbf.v4.new_markdown_cell(textwrap.dedent(text).strip("\n"))


def code(text: str):
    return nbf.v4.new_code_cell(textwrap.dedent(text).strip("\n"))


def obs(prompt: str = "Observations"):
    """Markdown placeholder for the student's analysis after running the section."""
    return md(f"""
> **{prompt}** — _fill in after running this section._
>
> _What worked, what didn't, what would you try next?_
""")


# ──────────────────────────────────────────────────────────────────────
# Build cells
# ──────────────────────────────────────────────────────────────────────

cells = []

# ────────────────────────── Title ──────────────────────────
cells.append(md("""
# Who Wants to Be a PoliMillionaire? — NLP Group Assignment 2025/26

| | |
|---|---|
| **Group members** | _Name 1_ — `email1@mail.polimi.it` · _Name 2_ — `email2@mail.polimi.it` · _Name 3_ — `email3@mail.polimi.it` |
| **Video link** | _paste URL here_ |
| **Default model** | Qwen 2.5 7B-Instruct (4-bit NF4) |
| **Default strategy** | Tiered: BaselineLLM → RAG → Ensemble · Maths → Agent |

This notebook is the **experimentation workbench** for the assignment.
The implementation lives in the `polimibot` Python package — this
notebook only configures, runs, compares, and saves results.

**How to use this notebook:**

1. Run **Section 0 — Setup** once per Colab session.
2. Edit knobs in **Section 1 — Configure** to choose your strategy.
3. Run **Section 2 — Run** to evaluate it and save a report.
4. Repeat 1→2 for as many configurations as you want.
5. Use **Section 3 — Compare** to read every saved report into a leaderboard.
6. **Section 4 — Save** writes consolidated CSVs for your final write-up.

Switching strategies should not require touching anything outside Section 1.
"""))

# ────────────────────────── Section 0 — Setup ──────────────────────────
cells.append(md("""
---

## 0. Setup

Install the package, import helpers, log in to the game server. Run this section once per Colab session — re-running it is harmless but slow.

> **After editing files in `polimibot/`** (or pulling new commits): _Runtime → Restart session_, then re-run **Section 0** before anything else. Even with `pip install -e .`, classes already imported in this kernel stay cached — restarting is the only reliable way to pick up the new code.
"""))

cells.append(md("""
### 0.0 Mount Google Drive (Colab only)

Two modes — controlled by the `clone` flag:

| `clone` | Code source | `data/` location |
|---------|-------------|-----------------|
| `True`  | Fresh git clone into `/content/PoliMillionaire/` (fast, not throttled) | Symlinked → Drive so run logs, gold set, and the RAG index survive session restarts |
| `False` | Work directly from the Drive copy | Native `data/` inside the Drive project folder |

When `clone=True`, a symlink `data/ → <drive_path>/data/` is created once and
reused on subsequent sessions — no data loss on kernel restart.
"""))

cells.append(code('''
# Mount Google Drive and set up the project. Two modes: clone or drive-direct.
import os
from google.colab import drive, userdata

drive.mount('/content/drive')
DRIVE_PROJECT = "/content/drive/MyDrive/Colab Notebooks/Polimillionaire"

clone = True  # ← flip to False to work entirely from the Drive copy

if clone:
    # Clone fresh code from GitHub into fast local filesystem.
    REPO_URL = f"https://{userdata.get('GITHUB_TOKEN')}@github.com/m-ebrahimzadeh/PoliMillionaire.git"
    if not os.path.exists('/content/PoliMillionaire'):
        print("Cloning repo…")
        !git clone {REPO_URL} /content/PoliMillionaire
    else:
        print("Repo already cloned — skipping clone.")
    %cd /content/PoliMillionaire

    # Symlink data/ → Drive so runs, gold sets, and the RAG index persist
    # across session restarts. Created once; reused on every subsequent run.
    import shutil as _shutil
    os.makedirs(f"{DRIVE_PROJECT}/data", exist_ok=True)
    if os.path.isdir('data') and not os.path.islink('data'):
        _shutil.rmtree('data')   # remove the empty placeholder from the fresh clone
    if not os.path.exists('data'):
        os.symlink(f"{DRIVE_PROJECT}/data", 'data')
    print(f'data/ → {os.readlink("data")}  (persistent on Drive)')
else:
    %cd "{DRIVE_PROJECT}"
    print(f'Working directly from Drive: {DRIVE_PROJECT}')
'''))

cells.append(md("### 0.1 Install"))

cells.append(code("""
# Install the project as an editable package, this cell does. Once per session, run it.
%pip install -q -e .
%pip install -q "transformers>=4.46,<4.50" "accelerate>=1.0,<1.5" "bitsandbytes>=0.45"
%pip install -q "faiss-cpu>=1.7" "sentence-transformers>=2.7" "wikipedia>=1.4"
%pip install -q matplotlib pandas
%pip install -q "openai-whisper>=20231117" "scipy>=1.11"
"""))

cells.append(md("### 0.2 Imports and helpers"))

cells.append(code("""
# All imports, here once. Reach further than this cell, the notebook should not.
from __future__ import annotations

import gc
import json
import os
import sys
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Optional

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

# Polimibot — the audited package.
from polimibot import (
    PATHS, RUNTIME, update_runtime, CATEGORIES, Category,
    GameQuestion, GameAdapter, AnswerOutcome, SessionRecord,
    Strategy, StrategyInput, StrategyOutput,
    GameResult, play_game,
    RunLogger, NullLogger, load_jsonl,
)
from polimibot.models.mock import MockLLM
from polimibot.prompts.templates import PromptStyle
from polimibot.strategies.llm_baseline import BaselineLLMStrategy
from polimibot.strategies.rag_strategy import RAGStrategy
from polimibot.strategies.tool_strategy import ToolStrategy
from polimibot.strategies.agent_strategy import AgentStrategy
from polimibot.strategies.ensemble_strategy import EnsembleStrategy
from polimibot.strategies.tiered_strategy import TieredStrategy, TierBreakpoints
from polimibot.tools.maths_tool import MathsTool
from polimibot.eval.evaluator import evaluate_strategy, EvalReport
from polimibot.eval.gold_set import GoldSet, load_gold_set, harvest_gold_set, save_gold_set
from polimibot.eval.wrong_set import WrongSet, load_wrong_set, harvest_wrong_set, save_wrong_set
from polimibot.eval.report_io import save_report, make_report_id
from polimibot.eval.calibration import calibration_from_runs, plot_calibration

# Make scripts/_session.py importable for the live-game session helper.
sys.path.insert(0, str(PATHS.project_root / 'scripts'))
from _session import play_session

PATHS.ensure()
print(f'project root : {PATHS.project_root}')
print(f'data/         : {PATHS.data_dir}')
print(f'API URL       : {RUNTIME.api_url}')
"""))

cells.append(code('''
# Helpers for VRAM hygiene and disk persistence — one place, defined they are.

def unload_llm(llm) -> None:
    """Free the LLM's GPU references. Call before loading another model."""
    if llm is None:
        return
    for attr in ('_model', '_tokenizer'):
        if hasattr(llm, attr):
            try:
                delattr(llm, attr)
            except Exception:
                pass


def clear_vram() -> None:
    """Release cached CUDA memory. Pair with unload_llm before any new load."""
    gc.collect()
    try:
        import torch
        if torch.cuda.is_available():
            torch.cuda.empty_cache()
            torch.cuda.synchronize()
    except ImportError:
        pass


RESULTS_DIR = PATHS.project_root / 'data' / 'results'


def save_results(df: pd.DataFrame, name: str) -> Path:
    """Persist a dataframe to data/results/{name}.csv. Crash-safety, this gives."""
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    path = RESULTS_DIR / f'{name}.csv'
    df.to_csv(path, index=False)
    print(f'Saved → {path}')
    return path


def latest_run_log(strategy_substring: str = '') -> Optional[Path]:
    """Most recent JSONL run log under data/runs/, optionally filtered by name."""
    runs = sorted(PATHS.runs_dir.glob('run_*.jsonl'))
    if strategy_substring:
        runs = [r for r in runs if strategy_substring in r.name]
    return runs[-1] if runs else None


print('helpers ready ✓')
'''))

cells.append(md("### 0.3 Log in (live games only — skip for offline eval)"))

cells.append(code('''
# Credentials, from Colab Secrets (userdata) or env vars we read.
# Set in Colab: Secrets panel → POLIMI_USER / POLIMI_PASS / GUARDIAN_API_KEY
#            or os.environ['POLIMI_USER'] = '...'  (fallback)
# Skip this cell entirely if only the offline gold set you intend to evaluate.
from millionaire_client import MillionaireClient

def _secret(name):
    """Read a secret from Colab userdata, falling back to the env var."""
    try:
        from google.colab import userdata as _userdata
        val = _userdata.get(name)
        if val:
            return val
    except Exception:
        pass
    return os.environ.get(name, '')

POLIMI_USER = _secret('POLIMI_USER')
POLIMI_PASS = _secret('POLIMI_PASS')

# Guardian key for the NEWS online source. Export it into the environment NOW
# so the strategy build (Section 1.3) — which reads os.environ at build time —
# picks it up no matter the import order. config.NEWS captures the key at
# import; setting it in a later cell only takes effect because the build
# threads os.environ['GUARDIAN_API_KEY'] in explicitly.
_guardian_key = _secret('GUARDIAN_API_KEY')
if _guardian_key:
    os.environ['GUARDIAN_API_KEY'] = _guardian_key
    print(f'GUARDIAN_API_KEY loaded (…{_guardian_key[-4:]}).')
else:
    print('GUARDIAN_API_KEY not found — NEWS online source will use Wikipedia/offline.')

client = None
if POLIMI_USER and POLIMI_PASS:
    client = MillionaireClient(RUNTIME.api_url)
    client.login(POLIMI_USER, POLIMI_PASS)
    print(f'Logged in as {POLIMI_USER} → {RUNTIME.api_url}')
else:
    print('No credentials found — live games disabled. Offline eval will still work.')
'''))

cells.append(md("""
### 0.4 Build / top-up the RAG knowledge index

Two phases so the **GPU-free harvest** and the **GPU embed/index** can run on different Colab runtimes:

- **0.4a — Harvest corpus (CPU-friendly).** Downloads the Wikipedia corpus and writes `data/cache/corpus.jsonl`. Pure network/CPU — run it on a **CPU runtime** to save GPU hours.
- **0.4b — Embed & index (GPU).** Loads `corpus.jsonl`, embeds with `bge-large-en-v1.5`, and writes the FAISS + BM25 index. Run it on a **GPU runtime**.

> **The CPU→GPU handoff is automatic** — no extra copying. Cell 1 symlinks `data/` to Drive (or you work directly from Drive), so `data/cache/corpus.jsonl` and the index already live on Drive and survive a runtime-type switch. To go CPU→GPU: run 0.4a, switch the runtime to GPU, re-run cells 0.1–0.3 + the knobs cell, then run 0.4b — it picks the corpus up from `data/cache` on its own.

| Knob | Meaning |
|---|---|
| `REBUILD_INDEX` | Set `True` to (re)build the index in 0.4b even if one already exists |
| `INDEX_REFETCH` | Set `True` to re-harvest in 0.4a even if `corpus.jsonl` exists |
| `INDEX_CATEGORIES` | `None` → all four categories; or e.g. `['history', 'science']` |
| `INDEX_CHUNK_SIZE` / `INDEX_OVERLAP` | Sentence-aware chunking parameters (default 300/50) |
| `INDEX_SKIP_BM25` | Set `True` to skip the BM25 sidecar (dense-only — disables hybrid retrieval) |
| `INDEX_LEGACY_SEEDS` | Use the hand-curated `TOPIC_SEEDS` (~95 titles) instead of the category-graph harvest |
| `INDEX_HARVEST_MAX_PER_CATEGORY` / `INDEX_HARVEST_MAX_DEPTH` | Harvester breadth / subcategory depth |
| `INDEX_HARVEST_WORKERS` | Concurrent extract batches (default 5 — polite + fast; raise for speed) |
| `INDEX_HARVEST_BATCH_SIZE` | Titles per extract request (default 20 — the anonymous MediaWiki cap) |
| `INDEX_GAP_QUEUE` | Path to `gap_titles.json` (scripts/mine_corpus_gaps.py) to back-fill, or `None` |
"""))

cells.append(code('''
# ─── RAG index knobs (shared by 0.4a + 0.4b). ────────────────────────
# RAG_INDEX_PATH / EMBEDDER_MODEL are also set in Section 1.1; mirrored here so
# 0.4 runs standalone — e.g. on a fresh GPU runtime after the CPU harvest.
RAG_INDEX_PATH     = PATHS.cache_dir / 'knowledge'
EMBEDDER_MODEL     = 'BAAI/bge-large-en-v1.5'   # must match Section 1.1; rebuild index after changing
REBUILD_INDEX      = False        # True  → (re)build index in 0.4b even if it exists
INDEX_REFETCH      = False        # True  → re-harvest in 0.4a even if corpus.jsonl exists
INDEX_CATEGORIES   = None         # None  → all four; or e.g. ['history', 'science']
INDEX_CHUNK_SIZE   = 300          # words per chunk (sentence-boundary-aware)
INDEX_OVERLAP      = 50           # word overlap between adjacent chunks
INDEX_SKIP_BM25    = False        # True  → skip BM25 sidecar (dense-only, faster)
INDEX_LEGACY_SEEDS = False        # True  → hand-curated TOPIC_SEEDS (~95 titles)
INDEX_HARVEST_MAX_PER_CATEGORY = 500   # cap per seed-category in the harvester
INDEX_HARVEST_MAX_DEPTH        = 0     # entity seeds: 0 = no recursion. Concept seeds always recurse 1 level.
INDEX_HARVEST_WORKERS          = 5     # concurrent extract batches (default 5 — polite + fast)
INDEX_HARVEST_BATCH_SIZE       = 20    # titles per extract request (MediaWiki cap is 20 for anonymous)
INDEX_GAP_QUEUE    = None         # path to gap_titles.json (scripts/mine_corpus_gaps.py) to back-fill, or None
# ─────────────────────────────────────────────────────────────────────
# data/ is symlinked to Drive by cell 1, so everything written under
# PATHS.cache_dir (corpus.jsonl, the index) is already durable across a
# runtime-type switch — no extra copy/restore step is needed.

print('RAG index knobs set. Run 0.4a (CPU harvest), then 0.4b (GPU embed/index).')
'''))

cells.append(md("#### 0.4a — Harvest corpus (CPU-friendly: Wikipedia → `corpus.jsonl`)"))

cells.append(code('''
# Phase A — download the corpus. Network/CPU only; no GPU needed. Safe to run on
# a CPU runtime to save GPU hours, then switch to GPU for 0.4b.
PATHS.ensure()
_corpus_path = PATHS.cache_dir / 'corpus.jsonl'

if _corpus_path.exists() and not INDEX_REFETCH:
    print(f'corpus.jsonl already present at {_corpus_path} — set INDEX_REFETCH=True to re-harvest.')
else:
    from polimibot.config import Category as _Category
    from polimibot.rag.corpus import (
        fetch_articles, fetch_articles_from_categories, fetch_articles_by_title,
        save_raw_corpus,
    )
    _cats = [_Category(c) for c in INDEX_CATEGORIES] if INDEX_CATEGORIES else None

    if INDEX_LEGACY_SEEDS:
        print('Fetching articles from Wikipedia (legacy hand-curated TOPIC_SEEDS)…')
        _articles = fetch_articles(categories=_cats, verbose=True)
    else:
        print('Fetching articles from Wikipedia (category-graph harvest, this takes several minutes)…')
        _articles = fetch_articles_from_categories(
            categories=_cats,
            cache_path=PATHS.cache_dir / 'harvested_titles.json',
            max_per_category=INDEX_HARVEST_MAX_PER_CATEGORY,
            max_depth=INDEX_HARVEST_MAX_DEPTH,
            harvest_workers=INDEX_HARVEST_WORKERS,
            batch_size=INDEX_HARVEST_BATCH_SIZE,
            checkpoint_path=_corpus_path,   # durable partial harvest — see corpus.py §8c
            verbose=True,
        )
    # Save the expensive harvest BEFORE the gap fetch, so nothing the gap phase
    # does can throw away the corpus. corpus.jsonl lives under data/cache, which
    # is symlinked to Drive (cell 1) — so this is already durable.
    save_raw_corpus(_articles, _corpus_path)

    # Log-mined gap back-fill: fetch the queued titles directly (see
    # scripts/mine_corpus_gaps.py). Skips titles already harvested.
    if INDEX_GAP_QUEUE:
        import json as _json
        from pathlib import Path as _Path
        _gap_path = _Path(INDEX_GAP_QUEUE)
        if _gap_path.is_file():
            _gap_raw = _json.loads(_gap_path.read_text(encoding='utf-8'))
            _gap_tbc = {}
            for _v, _ts in _gap_raw.items():
                try:
                    _gap_tbc[_Category(_v)] = list(_ts)
                except ValueError:
                    pass
            _gap_arts = fetch_articles_by_title(
                _gap_tbc, existing_titles={a.title for a in _articles}, verbose=True,
            )
            print(f'Gap queue added {len(_gap_arts)} articles')
            _articles = _articles + _gap_arts
            save_raw_corpus(_articles, _corpus_path)
        else:
            print(f'  ! INDEX_GAP_QUEUE {_gap_path} not found — skipping gap back-fill')

    if not _articles:
        raise RuntimeError('No articles fetched — check your network connection and INDEX_CATEGORIES.')
    print(f'\\nHarvest complete: {len(_articles)} articles → {_corpus_path}.')
    print('Switch to a GPU runtime (re-run 0.1-0.3 + the knobs cell) and run 0.4b.')
'''))

cells.append(md("#### 0.4b — Embed & index (GPU: `corpus.jsonl` → FAISS + BM25)"))

cells.append(code('''
# Phase B — embed + index. Run on a GPU runtime. corpus.jsonl is read from
# data/cache (Drive-backed via cell 1), so the 0.4a harvest is already here.
import time as _time
from dataclasses import replace as _replace
from polimibot.config import Category as _Category
from polimibot.rag.corpus import (
    load_raw_corpus, clean_wikipedia_text, CORPUS_VERSION, CLEANUP_VERSION,
)
from polimibot.rag.chunker import (
    CHUNKER_VERSION, EMBED_TEXT_VERSION,
    chunk_text as _chunk_text, embedding_text as _embedding_text,
)
from polimibot.rag.embedder import Embedder as _Embedder, EmbedderSpec as _EmbedderSpec
from polimibot.rag.index import FAISSIndex as _FAISSIndex
from polimibot.rag.bm25 import BM25Index as _BM25Index

PATHS.ensure()
_corpus_path = PATHS.cache_dir / 'corpus.jsonl'
_index_faiss = RAG_INDEX_PATH.with_suffix('.faiss')

if _index_faiss.exists() and not REBUILD_INDEX:
    print(f'Index already exists: {_index_faiss}')
    print('Set REBUILD_INDEX=True to force a rebuild.')
elif not _corpus_path.exists():
    raise RuntimeError(
        f'No corpus at {_corpus_path}. Run 0.4a first (the CPU harvest); it is '
        f'saved under data/cache, which is Drive-backed and survives a runtime switch.')
else:
    _cats = [_Category(c) for c in INDEX_CATEGORIES] if INDEX_CATEGORIES else None
    print(f'Loading corpus from {_corpus_path}…')
    _articles = load_raw_corpus(_corpus_path)
    if _cats:
        _articles = [a for a in _articles if a.category in _cats]
    _articles = [_replace(a, text=clean_wikipedia_text(a.text)) for a in _articles]
    if not _articles:
        raise RuntimeError('corpus.jsonl loaded but empty after the category filter.')

    # ── Step 1: chunk ────────────────────────────────────────────────
    print(f'\\nChunking {len(_articles)} articles (size={INDEX_CHUNK_SIZE}, overlap={INDEX_OVERLAP})…')
    _all_chunks = []
    for _art in _articles:
        _all_chunks.extend(_chunk_text(
            _art.text, source=_art.title,
            chunk_size=INDEX_CHUNK_SIZE, overlap=INDEX_OVERLAP,
            category=_art.category.value,
            url=_art.url,
            aliases=_art.aliases or None,
        ))
    print(f'  → {len(_all_chunks)} chunks '
          f'(avg {len(_all_chunks) // max(len(_articles), 1)} per article)')

    # ── Step 2: embed (GPU) ──────────────────────────────────────────
    print(f'\\nLoading embedding model ({EMBEDDER_MODEL})…')
    _spec = _EmbedderSpec(model_name=EMBEDDER_MODEL)
    _embedder = _Embedder(_spec)
    print(f'  dim={_embedder.dim}  batch={_spec.batch_size}')

    print('Embedding passages…')
    _t0 = _time.monotonic()
    _embeddings = _embedder.encode_passage([_embedding_text(c) for c in _all_chunks])
    print(f'  → done in {_time.monotonic() - _t0:.1f}s')

    # ── Step 3: FAISS index ──────────────────────────────────────────
    _idx = _FAISSIndex(dim=_embedder.dim)
    _idx.add(_all_chunks, _embeddings)
    _idx.save(RAG_INDEX_PATH, manifest={
        'embedder_model_name':     _spec.model_name,
        'embedder_dim':            _embedder.dim,
        'embedder_query_prefix':   _spec.query_prefix,
        'embedder_passage_prefix': _spec.passage_prefix,
        'normalize':               _spec.normalize,
        'chunk_size':              INDEX_CHUNK_SIZE,
        'chunk_overlap':           INDEX_OVERLAP,
        'chunker_version':         CHUNKER_VERSION,
        'embed_text_version':      EMBED_TEXT_VERSION,
        'corpus_version':          CORPUS_VERSION,
        'corpus_source':           'hand_curated' if INDEX_LEGACY_SEEDS else 'category_graph',
        'max_per_category':        INDEX_HARVEST_MAX_PER_CATEGORY,
        'max_depth':               INDEX_HARVEST_MAX_DEPTH,
        'n_articles':              len(_articles),
        'text_cleanup_version':    CLEANUP_VERSION,
        'categories':              sorted({a.category.value for a in _articles}),
    })
    print(f'\\n✓  FAISS index saved → {RAG_INDEX_PATH}.faiss')
    print(f'   {_idx.n_chunks} chunks | dim={_embedder.dim} | model={_spec.model_name}')

    # ── Step 4: BM25 sidecar ─────────────────────────────────────────
    if not INDEX_SKIP_BM25:
        print('\\nBuilding BM25 sidecar…')
        _t0 = _time.monotonic()
        _bm25 = _BM25Index(_all_chunks)
        _bm25.save(RAG_INDEX_PATH)
        print(f'   built in {_time.monotonic() - _t0:.1f}s  →  {RAG_INDEX_PATH}.bm25.jsonl')
    else:
        print('BM25 sidecar skipped (INDEX_SKIP_BM25=True).')

    # The index is written under data/cache (Drive-backed via cell 1), so it
    # already survives a runtime reset — no extra copy needed.
    print('\\nIndex build complete. Re-run Section 1.3 to attach the new index to the retriever.')
'''))

cells.append(obs("Setup observations"))

# ────────────────────────── Section 1 — Configure ──────────────────────────
cells.append(md("""
---

## 1. Configure

Pick a strategy by editing the knobs below. **One variable per concept** — change `MODEL_ID`, `PROMPT_STYLE`, `USE_RAG`, etc. independently. The next cell consumes these knobs and constructs a single `strategy` object.

**Common experiments:**
- _Plain LLM baseline_ → `USE_MOCK=False`, others default.
- _Few-shot vs zero-shot_ → toggle `PROMPT_STYLE`.
- _RAG ablation_ → flip `USE_RAG`.
- _Tools-only on maths_ → `USE_MATHS_TOOL=True`, `USE_RAG=False`.
- _Full tiered system_ → `USE_TIERED=True`.
- _Smoke test on CPU_ → `USE_MOCK=True`.
"""))

cells.append(md("### 1.1 Knobs"))

cells.append(code('''
# ─── Knobs. The only cell, edit you should, between experiments. ───

# Model
USE_MOCK            = False                              # CPU smoke-test mode (no GPU)
MODEL_ID            = 'Qwen/Qwen2.5-7B-Instruct'         # any HF causal LM
LOAD_IN_4BIT        = True                               # NF4 quantisation; False on CPU
TRUST_REMOTE_CODE   = False                              # set True for some Phi/DeepSeek/Yi releases

# Prompt
PROMPT_STYLE        = PromptStyle.ZERO_SHOT              # ZERO_SHOT, FEW_SHOT, *_COT
USE_SCORE_OPTIONS   = True                               # logit-scoring (False forces free generation)

# Generation budgets (only used when USE_SCORE_OPTIONS=False)
DIRECT_MAX_NEW_TOKENS = 16                               # ZERO_SHOT / FEW_SHOT — small, just "Answer: X"
COT_MAX_NEW_TOKENS    = 256                              # *_COT — room for the reasoning trace
STOP_STRINGS          = None                             # None → boxed-stops on CoT; or pass a list

# Per-question deadline (server gives 30s; 25s leaves margin for submit roundtrip)
HARD_CUTOFF_SECONDS = 25.0                               # strategy must return by this

# Speech mode
GAME_MODE           = "text"                               # "text" or "speech"
SPEECH_MODEL        = "small"                             # whisper model: "tiny", "base", "small", "medium"
SPEECH_DEVICE       = "cuda" if not USE_MOCK else "cpu"  # "cuda" or "cpu"

# Rebinds RUNTIME singleton, this does
update_runtime(hard_cutoff_seconds=HARD_CUTOFF_SECONDS)

# Strategy composition (highest USE_TIERED wins; otherwise stack from baseline up)
USE_RAG             = False                              # RAG over Wikipedia
USE_MATHS_TOOL      = False                              # deterministic arithmetic on maths Qs
USE_AGENT_FOR_MATHS = False                              # ReAct agent on maths Qs
USE_ENSEMBLE        = False                              # weighted-prob fusion of baseline + RAG
USE_TIERED          = False                              # full hybrid: tier-by-level with all the above
USE_CONFIDENCE_GATED  = True                            # primary → fallback-on-uncertainty
MARGIN_THRESHOLD      = 0.30                             # commit primary when (top1 - top2) prob ≥ this
# Categories that ALWAYS escalate to the RAG/live fallback, bypassing the margin
# gate. The base LLM cannot know dated NEWS articles (past its knowledge cutoff),
# so its confidence there is meaningless — retrieval is mandatory. [] → pure gate.
CONFGATE_ALWAYS_FALLBACK = ['news']                     # category values; e.g. [] or ['news']

# Tier knobs (only used when USE_TIERED=True)
TIER_EASY_MAX       = 5                                  # levels 1..5  → easy strategy
TIER_MEDIUM_MAX     = 10                                 # levels 6..10 → medium strategy
ESCALATION_THRESHOLD = None                              # e.g. 0.15 to escalate on low margin

# Category overrides — per-category specialist strategies that bypass the
# level-tier dispatch entirely (audit P1). The maths slot is filled below from
# USE_MATHS_TOOL / USE_AGENT_FOR_MATHS; add more entries here by name to
# enable a future category-specialist arm (e.g. ENTERTAINMENT → WikidataStrategy).
# Format: {'history': True, 'science': False, ...}.  Only categories set to True
# AND with a builder wired in 1.4 get an override; absent categories fall through
# to easy/medium/hard.
ENABLE_CATEGORY_OVERRIDES: dict[str, bool] = {
    'maths': True,    # follows USE_MATHS_TOOL / USE_AGENT_FOR_MATHS
}

# Retrieval (only used when USE_RAG / USE_ENSEMBLE / USE_TIERED)
RAG_K                  = 3                               # passages per question
RAG_INDEX_PATH         = PATHS.cache_dir / 'knowledge'   # build with scripts/build_rag_index.py
RAG_USE_CATEGORY_FILTER = True                           # restrict retrieval to inp.category chunks
RAG_USE_SCORE_OPTIONS  = True                            # logit-scoring (False = free generation; required for CoT/ELIMINATION)
RAG_MAX_PASSAGE_CHARS  = 800                             # per-passage truncation
RAG_MAX_TOTAL_CHARS    = 2400                            # joined context budget

# Reranker (cross-encoder over the dense pool — precision win, +~30 ms/query)
RAG_USE_RERANKER       = True                           # set True to load + use
RERANKER_MODEL         = 'BAAI/bge-reranker-base'       # lighter CE (278M), fp16; recalibrate RAG_MIN_SCORE_RERANK after swap
RERANK_OVERSEARCH      = 5                               # dense pool size = k × this

# Embedding model — single source of truth for index build (Section 0.4),
# index load (Section 1.3), and the IndexGrower embedder (Section 1.4).
# Switching models requires rebuilding the index (REBUILD_INDEX=True in 0.4).
# Prefixes auto-derive from the model name (BGE / E5 / symmetric MiniLM-style).
EMBEDDER_MODEL         = 'BAAI/bge-large-en-v1.5'        # 1024-dim dense, BGE query instruction (auto), fp16; rebuild index after changing (REBUILD_INDEX=True in 0.4)

# Hybrid + multi-query (lexical complement + per-option queries, both via RRF)
RAG_USE_HYBRID         = True                           # dense + BM25 fused per query
RAG_USE_MULTI_QUERY    = True                            # 1 question + 4 per-option queries (default on per audit §3)

# Path-aware min_score gates (audit §4): calibrate each threshold on its own
# score scale — never apply a cosine threshold to an RRF or reranker score.
RAG_MIN_SCORE          = None                            # dense-only cosine ∈ [-1,1]; e.g. 0.30
RAG_MIN_SCORE_RRF      = None                            # hybrid RRF ∈ ~0–0.03;      e.g. 0.010 (no-rerank path only)
RAG_MIN_SCORE_RERANK   = None                            # raw cross-encoder logit (pre-sigmoid, may be negative); calibrate via §2.8c  (≥1.0 skips offline → live-only)

# Live-search fallback + self-growing index (fires only when offline RAG is gated)
# When USE_LIVE_FALLBACK=True and the top offline retrieval score is below the
# min_score threshold, a real-time Wikipedia API query is fired instead of
# degrading to a bare-LLM prompt.  Confirmed-correct articles are added to
# the offline index permanently so future questions benefit.
USE_LIVE_FALLBACK      = True                           # True  → enable live Wikipedia fallback
LIVE_SEARCH_TIMEOUT    = 7.0                             # hard wall-clock limit per live query (s)
LIVE_MAX_ARTICLES      = 2                               # max Wikipedia articles per live query
LEARN_FROM_CORRECT     = True                            # grow the index from confirmed-correct answers
LIVE_USE_LLM_QUERY     = True                           # True → LLM distils question to 2-5 Wikipedia keywords before live search
MIN_LIVE_SCORE         = 0.20                            # absolute quality floor on live passages (cosine scale); None → keep all. Drops off-topic fallbacks instead of injecting them as context

# News online source (The Guardian) — NEWS category only. When True, NEWS
# questions route their gated live fallback through the Guardian (date + entity
# aware) instead of Wikipedia; NewsLiveSearch falls back to Wikipedia internally
# when the Guardian has nothing. Reads GUARDIAN_API_KEY from the environment;
# seed the offline News corpus with scripts/fetch_news_corpus.py --days 30 --build.
USE_NEWS_LIVE_SEARCH   = True                           # True → Guardian-backed live search for NEWS
NEWS_DATE_WINDOW_DAYS  = 2                              # ± days around the question's stated date (absorbs publish-date skew)
NEWS_MAX_ARTICLES      = 3                              # max Guardian articles per NEWS live query

# Eval
N_EVAL_QUESTIONS    = None                               # None = all gold items; int = first-N slice

print(f'mock={USE_MOCK}  model={MODEL_ID}  style={PROMPT_STYLE.value}')
print(f'game_mode={GAME_MODE}  speech_model={SPEECH_MODEL if GAME_MODE == "speech" else "n/a"}')
print(f'rag={USE_RAG}  maths_tool={USE_MATHS_TOOL}  agent={USE_AGENT_FOR_MATHS}')
print(f'ensemble={USE_ENSEMBLE}  tiered={USE_TIERED}')
print(f'hard_cutoff={HARD_CUTOFF_SECONDS}s  direct_max_new_tokens={DIRECT_MAX_NEW_TOKENS}  cot_max_new_tokens={COT_MAX_NEW_TOKENS}')
'''))

cells.append(md("### 1.2 Build the LLM (heavy — runs once per model change)"))

cells.append(code('''
# Try to reuse an already-loaded LLM. If the model id changed, free GPU memory first.
prev_model_id = globals().get('LOADED_MODEL_ID', None)
if 'llm' in globals() and prev_model_id != MODEL_ID:
    print(f'Model changed ({prev_model_id} → {MODEL_ID}). Unloading previous LLM…')
    unload_llm(llm)
    clear_vram()
    del llm

if 'llm' not in globals():
    if USE_MOCK:
        llm = MockLLM(name='mock', correctness=0.65)
        print('MockLLM ready (CPU, no GPU needed)')
    else:
        from polimibot.models.llm import LLM, LLMSpec
        spec = LLMSpec(
            model_id=MODEL_ID,
            load_in_4bit=LOAD_IN_4BIT,
            trust_remote_code=TRUST_REMOTE_CODE,
        )
        llm = LLM.load(spec)
    LOADED_MODEL_ID = MODEL_ID
else:
    print(f'Reusing already-loaded LLM: {llm.name}')
'''))

cells.append(md('#### 1.2.1 Build the Speech Transcriber (if "speech" game mode is on)'))

cells.append(code('''
transcriber = None
if GAME_MODE == "speech":
    from polimibot.models.speech import SpeechTranscriber, TranscriberSpec
    prev_speech = globals().get('LOADED_SPEECH_MODEL', None)
    if 'transcriber' not in globals() or transcriber is None or prev_speech != SPEECH_MODEL:
        transcriber = SpeechTranscriber.load(TranscriberSpec(
            model_name=SPEECH_MODEL,
            device=SPEECH_DEVICE,
        ))
        LOADED_SPEECH_MODEL = SPEECH_MODEL
        print(f'SpeechTranscriber ready: whisper-{SPEECH_MODEL} on {SPEECH_DEVICE}')
    else:
        print(f'Reusing already-loaded transcriber: {transcriber.name}')
else:
    print('Speech mode off — transcriber not loaded.')
'''))

cells.append(md("### 1.3 Build the retriever (only if RAG-related knobs are on)"))

cells.append(code('''
# Lazy retriever. Built only when at least one strategy needs it.
# Reranker is loaded once and attached to the retriever — heavy (~100 MB),
# so we cache it across re-runs of Section 1.3 the same way the LLM is cached.
need_retriever = USE_RAG or USE_ENSEMBLE or USE_TIERED or USE_CONFIDENCE_GATED

retriever = None
if need_retriever:
    # Cache the reranker model across Section 1 re-runs.
    prev_rer = globals().get('LOADED_RERANKER_MODEL', None)
    if RAG_USE_RERANKER and (
        'reranker_obj' not in globals() or prev_rer != RERANKER_MODEL
    ):
        from polimibot.rag.reranker import CrossEncoderReranker, RerankerSpec
        print(f'Loading cross-encoder reranker: {RERANKER_MODEL} …')
        reranker_obj = CrossEncoderReranker.load(
            RerankerSpec(model_name=RERANKER_MODEL)
        )
        LOADED_RERANKER_MODEL = RERANKER_MODEL
    elif not RAG_USE_RERANKER:
        reranker_obj = None
    else:
        print(f'Reusing already-loaded reranker: {reranker_obj.name}')

    if USE_MOCK:
        # Null retriever for offline smoke tests, this is.
        class _NullRetriever:
            n_chunks = 0
            has_reranker = False
            has_bm25 = False
            def retrieve(self, q, k=3, *, category=None, rerank=False,
                         rerank_oversearch=None, hybrid=False):
                return []
        retriever = _NullRetriever()
        print('NullRetriever (mock mode — no FAISS index needed)')
    else:
        if not RAG_INDEX_PATH.with_suffix('.faiss').exists():
            raise FileNotFoundError(
                f'RAG index missing: {RAG_INDEX_PATH}.faiss\\n'
                'Build it first:\\n'
                '  python scripts/build_rag_index.py'
            )
        from polimibot.rag.retriever import Retriever
        from polimibot.rag.embedder import EmbedderSpec
        retriever = Retriever.from_saved(
            RAG_INDEX_PATH, embedder_spec=EmbedderSpec(model_name=EMBEDDER_MODEL),
        )
        if reranker_obj is not None:
            retriever._reranker = reranker_obj   # late-attach
        # Late-attach BM25 if its sidecar exists (built by --no-bm25=False).
        if RAG_USE_HYBRID:
            from polimibot.rag.bm25 import BM25Index
            bm25_path = RAG_INDEX_PATH.with_suffix('.bm25.jsonl')
            if not bm25_path.exists():
                raise FileNotFoundError(
                    f'BM25 sidecar missing: {bm25_path}\\n'
                    'Rebuild the index without --no-bm25:\\n'
                    '  python scripts/build_rag_index.py'
                )
            retriever._bm25 = BM25Index.load(RAG_INDEX_PATH)
        rer_tag = f' + reranker {reranker_obj.name}' if reranker_obj else ''
        bm25_tag = f' + BM25 ({retriever._bm25.n_chunks} chunks)' if retriever.has_bm25 else ''
        print(f'Retriever ready: {retriever.n_chunks} chunks indexed{rer_tag}{bm25_tag}')
else:
    print('Retriever not needed for this configuration.')
    reranker_obj = None
'''))

cells.append(md("### 1.4 Compose the strategy"))

cells.append(code('''
# Strategy factory. Knobs in, one Strategy out — the only place composition lives.

baseline = BaselineLLMStrategy(
    llm,
    style=PROMPT_STYLE,
    use_score_options=USE_SCORE_OPTIONS,
    direct_max_new_tokens=DIRECT_MAX_NEW_TOKENS,
    cot_max_new_tokens=COT_MAX_NEW_TOKENS,
    stop_strings=STOP_STRINGS,
)

# ── IndexGrower (self-growing index) ────────────────────────────────────
# Built only when USE_LIVE_FALLBACK=True AND a real (non-mock) retriever is
# available.  The grower buffers live-fetched articles during gameplay; the
# runner confirms and persists them at session end.  For offline eval it is
# unused — the grower.confirm() / flush() calls in runner.py are no-ops when
# grower is None.
_grower = None
if USE_LIVE_FALLBACK and need_retriever and not USE_MOCK:
    from polimibot.rag.index_grower import IndexGrower
    from polimibot.rag.embedder import Embedder
    _embedder_for_grower = Embedder(retriever._embedder.spec)   # exact spec used to load retriever
    _grower = IndexGrower(
        retriever, _embedder_for_grower, RAG_INDEX_PATH,
        corpus_path=PATHS.cache_dir / 'corpus.jsonl',
    )
    print(f'IndexGrower ready — will learn from {_grower.n_learned} confirmed articles so far')
elif USE_LIVE_FALLBACK:
    print('IndexGrower: skipped (USE_MOCK=True or no retriever) — live search context-only mode')

# ── News online source (Guardian) for the NEWS category ─────────────────
# Built only when USE_NEWS_LIVE_SEARCH=True and a real (non-mock) retriever is
# in play. NEWS questions route their gated live fallback here (date + entity
# aware); every other category keeps using the Wikipedia LiveSearchFallback.
# NewsLiveSearch falls back to Wikipedia internally when the Guardian returns
# nothing or no GUARDIAN_API_KEY is set.
_news_search = None
if USE_NEWS_LIVE_SEARCH and need_retriever and not USE_MOCK:
    from polimibot.rag.news_search import NewsLiveSearch
    from polimibot.config import update_news
    # Read the key from the environment HERE (build time), not at import time —
    # this is what lets the Colab secret cell take effect even though it runs
    # after `import polimibot`.
    _news_cfg = update_news(
        guardian_api_key=os.environ.get('GUARDIAN_API_KEY', '').strip(),
        date_window_days=NEWS_DATE_WINDOW_DAYS,
        max_articles=NEWS_MAX_ARTICLES,
        timeout_seconds=LIVE_SEARCH_TIMEOUT,
    )
    _news_search = NewsLiveSearch(_news_cfg)
    _key_state = (f'GUARDIAN_API_KEY set (…{_news_cfg.guardian_api_key[-4:]})'
                  if _news_cfg.guardian_api_key
                  else 'no key — Wikipedia fallback only')
    print(f'NewsLiveSearch ready for NEWS — {_key_state}')
elif USE_NEWS_LIVE_SEARCH:
    print('NewsLiveSearch: skipped (USE_MOCK=True or no retriever) — NEWS uses Wikipedia fallback')

# Shared RAGStrategy kwargs — assemble once so all three call sites stay in sync.
_rag_kwargs = dict(
    k=RAG_K,
    style=PROMPT_STYLE,
    use_score_options=RAG_USE_SCORE_OPTIONS,
    use_category_filter=RAG_USE_CATEGORY_FILTER,
    use_reranker=RAG_USE_RERANKER,
    use_hybrid=RAG_USE_HYBRID,
    use_multi_query=RAG_USE_MULTI_QUERY,
    rerank_oversearch=RERANK_OVERSEARCH,
    # Path-aware min_score gates (audit §4): each threshold is calibrated
    # for its own score scale — dense cosine, RRF, or cross-encoder logit.
    min_score=RAG_MIN_SCORE,
    min_score_rrf=RAG_MIN_SCORE_RRF,
    min_score_rerank=RAG_MIN_SCORE_RERANK,
    min_live_score=MIN_LIVE_SCORE,
    max_passage_chars=RAG_MAX_PASSAGE_CHARS,
    max_total_chars=RAG_MAX_TOTAL_CHARS,
    # Live-search fallback — fires only when offline retrieval is gated.
    use_live_fallback=USE_LIVE_FALLBACK,
    live_search_timeout=LIVE_SEARCH_TIMEOUT,
    live_max_articles=LIVE_MAX_ARTICLES,
    index_grower=_grower if LEARN_FROM_CORRECT else None,
    live_use_llm_query=LIVE_USE_LLM_QUERY,
    # NEWS-only Guardian source; None → NEWS uses the Wikipedia fallback.
    news_search=_news_search,
)

if USE_CONFIDENCE_GATED:
    from polimibot.strategies.confidence_gated_strategy import ConfidenceGatedStrategy
    from polimibot.config import Category as _Category_cg
    # Primary: bare LLM, logit-scored. Uses the `baseline` object built at
    # the top of this cell. No RAG, no live — fast and well-calibrated.
    # Fallback: live-Wikipedia-only RAG (offline auto-skipped when
    # RAG_MIN_SCORE_RERANK >= 1.0). Fires when primary's margin is below
    # MARGIN_THRESHOLD, or always for CONFGATE_ALWAYS_FALLBACK categories
    # (NEWS) where the model's confidence is not a trustworthy signal.
    fallback_rag = RAGStrategy(llm, retriever, **_rag_kwargs)
    strategy = ConfidenceGatedStrategy(
        primary=baseline,
        fallback=fallback_rag,
        margin_threshold=MARGIN_THRESHOLD,
        always_fallback_categories=frozenset(
            _Category_cg(c) for c in CONFGATE_ALWAYS_FALLBACK
        ),
    )
elif USE_TIERED:
    from polimibot.config import Category as _Category_tiered
    rag_arm    = RAGStrategy(llm, retriever, **_rag_kwargs)
    ensemble   = EnsembleStrategy([baseline, rag_arm], weights=[1.0, 1.2])
    # Build per-category override arms. None of these have to be present;
    # absent categories fall through to the level-tier dispatch.
    _cat_overrides: dict = {}
    if ENABLE_CATEGORY_OVERRIDES.get('maths', False):
        _maths_arm = (
            AgentStrategy(llm, max_iterations=3) if USE_AGENT_FOR_MATHS
            else (ToolStrategy([MathsTool()], fallback=baseline) if USE_MATHS_TOOL else None)
        )
        if _maths_arm is not None:
            _cat_overrides[_Category_tiered.MATHS] = _maths_arm
    # Future category-specialist arms slot in here, gated by their own
    # ENABLE_CATEGORY_OVERRIDES flag. Example (when Wikidata is added):
    #   if ENABLE_CATEGORY_OVERRIDES.get('entertainment', False):
    #       _cat_overrides[_Category_tiered.ENTERTAINMENT] = WikidataStrategy(llm, fallback=rag_arm)
    strategy = TieredStrategy(
        easy=baseline, medium=rag_arm, hard=ensemble,
        breakpoints=TierBreakpoints(
            easy_max_level=TIER_EASY_MAX,
            medium_max_level=TIER_MEDIUM_MAX,
        ),
        category_overrides=_cat_overrides,
        escalation_threshold=ESCALATION_THRESHOLD,
    )
elif USE_ENSEMBLE:
    rag_arm  = RAGStrategy(llm, retriever, **_rag_kwargs)
    strategy = EnsembleStrategy([baseline, rag_arm], weights=[1.0, 1.2])
elif USE_AGENT_FOR_MATHS:
    strategy = AgentStrategy(llm, max_iterations=3)
elif USE_MATHS_TOOL:
    strategy = ToolStrategy([MathsTool()], fallback=baseline)
elif USE_RAG:
    strategy = RAGStrategy(llm, retriever, **_rag_kwargs)
else:
    strategy = baseline

# report_id is computed here (Section 1) — Section 2 saves under it, Section 2.6
# uses it as the live-game run_id. Defining it after the strategy means a student
# can run live games (2.6) without first running offline eval (2.2 / 2.3).
report_id  = make_report_id(strategy, MODEL_ID, PROMPT_STYLE, mock=USE_MOCK)

print(f'Strategy:\\n  {strategy.name}')
print(f'report_id: {report_id}')
if USE_LIVE_FALLBACK:
    print(f'Live-search fallback: ENABLED (timeout={LIVE_SEARCH_TIMEOUT}s, max_articles={LIVE_MAX_ARTICLES})')
    print(f'Index growing: {"ENABLED (learn from correct answers)" if LEARN_FROM_CORRECT else "DISABLED"}')
if _news_search is not None:
    print(f'NEWS online source: Guardian (±{NEWS_DATE_WINDOW_DAYS}d window, max_articles={NEWS_MAX_ARTICLES})')
'''))

cells.append(md("""
### 1.5 Strategy smoke-test — inspect the prompt & output

Run a single question through the strategy to verify the pipeline end-to-end and
see exactly what is delivered to the model.

1. **Bare prompt** — the raw messages list (system + user turn) that the LLM receives
   when no RAG context is available.
2. **RAG-augmented prompt** — the same structure with retrieved passages injected
   (only shown when the retriever is active).
3. **Strategy output** — the chosen answer, confidence score, rationale, and any
   internal extras (probability margins, retrieval hits, etc.).
"""))

cells.append(code('''
# ─── Single-question smoke test. See what the model actually receives. ───
from polimibot.prompts.templates import build_messages, build_messages_with_context

_test_question = (
    'A farmer wants to know whether a new fertilizer has increased the mean '
    'weight of his apples. With the old fertilizer, the mean weight was 4.0 '
    'ounces per apple. The farmer decides to test H0: μ = 4.0 ounces versus '
    'Ha : μ > 4.0 ounces, at a 5 percent level of significance, where μ = '
    'the mean weight of apples using the new fertilizer. The weights of apples '
    'are approximately normally distributed. The farmer takes a random sample '
    'of 16 apples and computes a mean of 4.3 ounces and a standard deviation '
    'of 0.6 ounces. Which of the following gives the p-value for this test?'
)
_test_options = (
    'P(Z > 2)',
    'P(t > 2) with 15 degrees of freedom',
    'P(t < 2) with 15 degrees of freedom',
    'P(Z < 2)',
)

# ── 1. Bare prompt (no RAG context) ──────────────────────────────────
_bare_msgs = build_messages(_test_question, _test_options,
                            category=None, style=PROMPT_STYLE)
print('═' * 60)
print('BARE PROMPT (no RAG context)')
print('═' * 60)
for _m in _bare_msgs:
    print(f"\\n[{_m['role'].upper()}]")
    print(_m['content'])

# ── 2. RAG-augmented prompt (only when retriever is active) ───────────
if retriever is not None and retriever.n_chunks > 0:
    _hits = retriever.retrieve(_test_question, k=RAG_K)
    _ctx  = '\\n---\\n'.join(chunk.text[:RAG_MAX_PASSAGE_CHARS] for chunk, _score in _hits)
    _rag_msgs = build_messages_with_context(
        _test_question, _test_options, _ctx,
        category=None, style=PROMPT_STYLE,
    )
    print('\\n' + '═' * 60)
    print(f'RAG-AUGMENTED PROMPT  ({len(_hits)} passage(s) retrieved)')
    print('═' * 60)
    for _m in _rag_msgs:
        print(f"\\n[{_m['role'].upper()}]")
        print(_m['content'])
else:
    print('\\n(Retriever not active — RAG-augmented prompt not shown.)')

# ── 3. Full strategy output ───────────────────────────────────────────
print('\\n' + '═' * 60)
print('STRATEGY OUTPUT')
print('═' * 60)
_out = strategy.answer(StrategyInput(
    question=_test_question, options=_test_options, level=1))
print('rationale :', _out.rationale)
print('chosen    :', _out.chosen_index, '  (A=0, B=1, C=2, D=3)')
print('confidence:', f'{_out.confidence:.2%}')
if _out.extras:
    print('extras    :', _out.extras)
'''))

cells.append(obs("Configuration observations"))

# ────────────────────────── Section 2 — Run ──────────────────────────
cells.append(md("""
---

## 2. Run

Two ways to exercise the strategy:

- **Offline evaluation** on the frozen gold set (fast, deterministic, no network).
- **Live games** against the assignment server (slow, costs API calls, requires login).

The notebook prefers offline. Run live only when you want a fresh row of run-log data, or when you need to top up the gold set.
"""))

cells.append(md("""
### 2.0 Build / top-up the gold set from run logs

Harvests confirmed-correct items from all JSONL run logs in `data/runs/` and
merges them into `data/eval/gold_set.jsonl`. Run this after live games so the
freshly confirmed answers are available for offline evaluation.

Safe to re-run at any time — items already present are not duplicated.
"""))

cells.append(code('''
# Top up the gold set with items confirmed during live games, this cell does.
gold_path = PATHS.eval_dir / 'gold_set.jsonl'
items = harvest_gold_set(PATHS.runs_dir)
# To filter by model, use keyword args — substring-match on the manifest's extra.model_id / extra.model:
# items = harvest_gold_set(PATHS.runs_dir, exclude_models=['qwen'])
# items = harvest_gold_set(PATHS.runs_dir, include_models=['gpt'])
# items = harvest_gold_set(PATHS.runs_dir, run_filter=lambda m: m['extra'].get('seed') == 42)

print(f'Harvested {len(items)} gold items from {PATHS.runs_dir}')
for cid, info in CATEGORIES.items():
    n = sum(1 for it in items if it.competition_id == cid)
    print(f'  {info.display_name:<35} {n:>4} items')

if items:
    save_gold_set(items, gold_path)
    print(f'\\n✓ Gold set saved → {gold_path}')
else:
    print('\\nNo items confirmed yet — play more live games first.')
'''))

cells.append(md("""
### 2.0.1 Build / top-up the wrong set from run logs

Symmetrical companion to the gold set. Harvests questions the bot answered
*incorrectly* from all JSONL run logs in `data/runs/` and saves them to
`data/eval/wrong_set.jsonl`.

**Why collect wrong questions?**
- Error analysis by category and level — spot systematic weaknesses.
- Understand *how* the bot fails (which distractors fool it most).
- Targeted re-evaluation after strategy changes.

The `correct_index` field is filled in when the correct answer can be
recovered from the same run logs (direct confirmation or elimination);
otherwise it is `-1`. Safe to re-run at any time.
"""))

cells.append(code('''
# Top up the wrong set with incorrectly-answered items from live games.
wrong_path = PATHS.eval_dir / 'wrong_set.jsonl'
wrong_items = harvest_wrong_set(PATHS.runs_dir)

print(f'Harvested {len(wrong_items)} wrong items from {PATHS.runs_dir}')
for cid, info in CATEGORIES.items():
    n = sum(1 for it in wrong_items if it.competition_id == cid)
    print(f'  {info.display_name:<35} {n:>4} items')

if wrong_items:
    n_known = sum(1 for it in wrong_items if it.correct_index >= 0)
    print(f'\\nCorrect answer recovered: {n_known}/{len(wrong_items)} '
          f'({n_known / len(wrong_items):.1%})')
    save_wrong_set(wrong_items, wrong_path)
    print(f'\\n✓ Wrong set saved → {wrong_path}')
else:
    print('\\nNo wrong items found yet — play more live games first.')
'''))

cells.append(md("""
### 2.0.2 Load and inspect the wrong set

`WrongSet` is a chainable view over the wrong items — every filter / sampler /
splitter returns a new `WrongSet`, mirroring the `GoldSet` API exactly.

**Recipes:**

```python
wrongs        = WrongSet.load(wrong_path)          # everything
maths         = wrongs.filter_category(Category.MATHS)
hard          = wrongs.filter_level(min_level=11)
known         = wrongs.filter_known_correct()      # correct answer is known
balanced      = wrongs.take_per_level(3, seed=0)   # ≤3 per level 1..15
train, test   = wrongs.split(0.8, seed=42)
```
"""))

cells.append(code('''
# Wrong set, from data/eval/wrong_set.jsonl we load.
wrong_path = PATHS.eval_dir / 'wrong_set.jsonl'

if not wrong_path.exists():
    print(
        f'Wrong set not found at {wrong_path}.\\n'
        'Build it by running the cell above, or:\\n'
        '  python scripts/build_wrong_set.py\\n'
        'Or play games first to populate data/runs/, then re-build.'
    )
else:
    wrongs = WrongSet.load(wrong_path)
    print(f'Wrong set: {len(wrongs)} items')
    wrongs.print_stats()

    # ── Per-category error bar chart ─────────────────────────────────────
    by_cat = wrongs.counts_by_category()
    if by_cat:
        cats = sorted(by_cat)
        counts = [by_cat[c] for c in cats]

        fig, ax = plt.subplots(figsize=(7, 3.5))
        bars = ax.bar(cats, counts, color=\'salmon\', edgecolor=\'white\')
        for bar, cnt in zip(bars, counts):
            ax.text(bar.get_x() + bar.get_width() / 2,
                    bar.get_height() + 0.1,
                    str(cnt), ha=\'center\', fontsize=9)
        ax.set_ylabel(\'Wrong answers\')
        ax.set_title(f\'Wrong answers by category  (total={len(wrongs)})\')
        plt.tight_layout()
        plt.show()

    # ── Cross-reference: questions both wrong and in the gold set ─────────
    if wrong_path.exists() and gold_path.exists():
        gold_keys  = {(g.competition_id, g.question_text) for g in full}
        overlap    = [w for w in wrongs if (w.competition_id, w.question_text) in gold_keys]
        print(f\'\\nOverlap (in both gold + wrong set): {len(overlap)} items\')
        if overlap:
            print(\'These were answered correctly on at least one run and wrongly on another.\')
'''))

cells.append(md("""
### 2.1 Load the gold set

`GoldSet` is a chainable view over the gold items — every filter / sampler / splitter returns a new `GoldSet`, so you can branch experiments freely.

**Recipes:**

```python
full          = GoldSet.load(gold_path)          # everything
maths         = full.filter_category(Category.MATHS)
maths_hard    = maths.filter_level(min_level=11)
balanced      = full.take_per_level(3, seed=0)   # ≤3 per level 1..15
random_pilot  = full.sample(20, seed=42)         # 20 random items
train, test   = full.split(0.8, seed=42)         # 80/20 shuffled split
holdout       = full - test                      # set difference by question identity
```

Drop the resulting `GoldSet` straight into `evaluate_strategy(strategy, gold, ...)` — it's iterable and `__len__`-able.
"""))

cells.append(code('''
# Gold set, from data/eval/gold_set.jsonl we load. Build it first if missing.
gold_path = PATHS.eval_dir / 'gold_set.jsonl'

if not gold_path.exists():
    raise FileNotFoundError(
        f'Gold set not found at {gold_path}.\\n'
        'Build it with one of:\\n'
        '  python scripts/build_gold_set.py     (mines existing run logs)\\n'
        'Or play games first to populate data/runs/, then re-build.'
    )

full = GoldSet.load(gold_path)
print(f'Full gold set: {len(full)} items')
full.print_stats()
'''))

cells.append(md("""
### 2.1.1 Choose your eval subset

Pick **one** of the recipes below — or write your own — and assign the result to `gold`. Section 2.2 evaluates against whatever's in `gold`.
"""))

cells.append(code('''
# ─── Eval-subset selector. Edit me. ──────────────────────────────────

# (a) Default: full set, optionally capped to first N items.
gold = full if N_EVAL_QUESTIONS is None else full.take(N_EVAL_QUESTIONS)

# (b) Single-category — uncomment one.
# gold = full.filter_category(Category.MATHS)
# gold = full.filter_category(Category.SCIENCE)
# gold = full.filter_category(Category.HISTORY)
# gold = full.filter_category(Category.ENTERTAINMENT)

# (c) Difficulty slice.
# gold = full.filter_level(min_level=1,  max_level=5)      # easy tier
# gold = full.filter_level(min_level=6,  max_level=10)     # medium tier
# gold = full.filter_level(min_level=11, max_level=15)     # hard tier

# (d) Difficulty-balanced pilot — at most N per level.
# gold = full.take_per_level(3, seed=0)

# (e) Category-balanced pilot — at most N per category.
# gold = full.take_per_category(10, seed=0)

# (f) Random sample — reproducible with the seed.
# gold = full.sample(50, seed=42)

# (g) Train / held-out test.
# train, test = full.split(0.8, seed=42)
# gold = test     # evaluate on held-out only

# (h) Chain freely.
# gold = full.filter_category(Category.MATHS).filter_level(min_level=8).take_per_level(2, seed=0)

print(f'Eval subset: {len(gold)} items')
gold.print_stats()
'''))

cells.append(md("### 2.2 Evaluate the strategy (offline)"))

cells.append(code('''
# Warm up once, evaluate, never warm again — this is the controlled comparison.
strategy.warm_up()

t0 = time.monotonic()
report: EvalReport = evaluate_strategy(strategy, gold, verbose=True)
print(f'\\nTotal eval time: {time.monotonic() - t0:.1f}s')

report.print_summary()
'''))

cells.append(md("""
### 2.2.1 Retrieval health dashboard (optional)

Prints aggregate pipeline diagnostics across the entire eval run:

| Section | What you see |
|---|---|
| **RETRIEVAL GATING** | How often offline retrieval was gated by `min_score`; mean top-score for correct vs wrong questions; gated-vs-ungated accuracy; per-category gate rates |
| **LIVE SEARCH** | How many times live Wikipedia fallback fired; success rate; p50/p95 latency |
| **TIER ROUTING** | Which tier (easy/medium/hard/maths) handled each question; escalation rate |
| **ENSEMBLE ARMS** | Which arm agreed with the final decision most often |

This cell is a no-op for non-RAG strategies — it degrades gracefully to a single notice line.
"""))

cells.append(code('''
# Retrieval health dashboard — shows what happened inside the pipeline.
# Works with any strategy; silent no-op when extras are absent (e.g. BaselineLLMStrategy).
from polimibot.observability import retrieval_dashboard

retrieval_dashboard(report)
'''))

cells.append(md("""
### 2.2.2 Per-question trace viewer (optional)

Drill into individual questions to see the **full pipeline trace** — exactly what was
retrieved, whether gating fired, what the live search returned, which tier/arm handled
the question, and what confidence margin was produced.

**Recipes:**

```python
# Show the 3 most confidently wrong answers (highest confidence, but wrong)
wrong = [s for s in report.samples if not s.correct]
for i, s in enumerate(sorted(wrong, key=lambda x: -x.confidence)[:3]):
    show_trace(s, idx=i)

# Show the 3 lowest-confidence wrong answers
for i, s in enumerate(sorted(wrong, key=lambda x: x.confidence)[:3]):
    show_trace(s, idx=i)

# Show a specific question by index
show_trace(report.samples[42], idx=42)

# Show all questions where live search fired
live_fired = [s for s in report.samples if s.extras.get('live_search_fired')]
for i, s in enumerate(live_fired):
    show_trace(s, idx=i)

# Show all gated questions that were answered wrong
gated_wrong = [s for s in report.samples
               if s.extras.get('gated_by_min_score') and not s.correct]
for i, s in enumerate(gated_wrong[:5]):
    show_trace(s, idx=i)
```
"""))

cells.append(code('''
# Per-question trace viewer. Edit the filter below to focus on the cases you care about.
from polimibot.observability import show_trace

# ── Choose which questions to inspect. ───────────────────────────────────
# (a) Default: 5 most confidently wrong answers
wrong = [s for s in report.samples if not s.correct]
_to_show = sorted(wrong, key=lambda x: -x.confidence)[:5]

# (b) All questions where gating fired and the answer was wrong — uncomment to use:
# _to_show = [s for s in report.samples
#             if s.extras.get('gated_by_min_score') and not s.correct]

# (c) All questions where live search fired — uncomment to use:
# _to_show = [s for s in report.samples if s.extras.get('live_search_fired')]

# (d) A specific index — uncomment to use:
# _to_show = [report.samples[0]]
# ─────────────────────────────────────────────────────────────────────────

if not _to_show:
    print('No matching samples to show.')
else:
    print(f'Showing {len(_to_show)} question(s):')
    for _i, _s in enumerate(_to_show):
        show_trace(_s, idx=_i)
'''))

cells.append(md("### 2.3 Save the report"))

cells.append(code('''
# Persistent on disk, every report is. Section 3, the leaderboard, reads these.
# report_id was defined in Section 1.4 — see comment there.
report_path = save_report(report, name=report_id, eval_dir=PATHS.eval_dir)
print(f'Report saved as: {report_id}')
'''))

cells.append(md("### 2.4 Per-category accuracy plot"))

cells.append(code('''
# Per-category accuracy heatmap, prefer over numbers we do.
cats = sorted(report.by_category.keys())
accs = [report.by_category[c].accuracy for c in cats]
ns   = [report.by_category[c].n        for c in cats]

fig, ax = plt.subplots(figsize=(7, 3.5))
bars = ax.bar(cats, accs, color='steelblue', edgecolor='white')
for bar, n in zip(bars, ns):
    ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.01,
            f'n={n}', ha='center', fontsize=9)
ax.set_ylim(0, 1.05)
ax.set_ylabel('Accuracy')
ax.set_title(f'{strategy.name}\\noverall acc = {report.accuracy:.1%}')
ax.axhline(0.25, linestyle='--', color='grey', linewidth=1, label='random baseline (0.25)')
ax.legend(loc='lower right')
plt.tight_layout()
plt.show()

# Save the per-category numbers, separately for write-up tables.
df_cat = pd.DataFrame(
    [(c, s.n, s.correct, s.accuracy, s.ece) for c, s in report.by_category.items()],
    columns=['category', 'n', 'correct', 'accuracy', 'ece'],
)
save_results(df_cat, f'percategory__{report_id}')
'''))

cells.append(md("### 2.5 Reliability diagram (calibration)"))

cells.append(code('''
# Calibration plot. Confidence vs accuracy buckets, overconfidence diagnose this does.
# Source: each EvalSample's confidence/correct, written by evaluate_strategy in-memory.
confidences = [s.confidence for s in report.samples]
corrects    = [s.correct    for s in report.samples]

from polimibot.eval.calibration import compute_calibration
cal = compute_calibration(confidences, corrects, n_bins=10)
plot_calibration(cal, title=f'Reliability — {strategy.name}',
                 output_path=PATHS.eval_dir / f'calibration__{report_id}.png')
'''))

cells.append(md("### 2.6 Live games (optional)"))

cells.append(code('''
# Live play, opt-in this is. Skip if no client. Each game writes a JSONL run log.
RUN_LIVE_GAMES         = False                         # ← flip to True to play
GAMES_PER_COMPETITION  = 1
COMPS_TO_PLAY          = list(CATEGORIES.keys())       # [0, 2, 3] for specific categories
# 0 = entertainment, 1= history, 2=science, 3=math, 4=philosohpy, 5=current_news

if RUN_LIVE_GAMES:
    if client is None:
        raise RuntimeError('Set POLIMI_USER / POLIMI_PASS in env, then re-run Section 0.3.')
    results = play_session(
        client,
        competition_ids=COMPS_TO_PLAY,
        strategy=strategy,
        games_per_competition=GAMES_PER_COMPETITION,
        run_id=report_id,
        verbose=True,
    )

    if GAME_MODE != "speech":
        df_live = pd.DataFrame([{
            'competition'   : r.competition_name,
            'final_level'   : r.final_level,
            'earned'        : r.earned_amount,
            'accuracy'      : r.accuracy,
            'elapsed_s'     : r.elapsed_seconds,
            'n_questions'   : r.n_questions,
        } for r in results])

        from datetime import datetime
        _ts  = datetime.now().strftime('%Y%m%d_%H%M%S')

        save_results(df_live, f'live__{report_id}__{_ts}')
        df_live
    else:
        print(f'GAME_MODE = {GAME_MODE}. Results not saved to the gold set.')
else:
    print('RUN_LIVE_GAMES=False — skipping live play.')
'''))

cells.append(md("""
### 2.6.1 Post-game retrieval summary (live games only)

After each live game session, this cell prints a full diagnostic breakdown:
context source used (offline RAG / live search / no context), per-category accuracy,
difficulty curve, live-search efficiency, confidence calibration, and worst misses.

Requires `RUN_LIVE_GAMES = True` in the cell above.
"""))

cells.append(code('''
# Post-game retrieval summary — one report per competition played.
# Needs RUN_LIVE_GAMES=True above; skips gracefully otherwise.
from polimibot.observability import print_game_summary

if RUN_LIVE_GAMES and results:
    for r in results:
        print_game_summary(r)
else:
    print('Skipping post-game summary (RUN_LIVE_GAMES=False or no results).')
'''))

cells.append(md("""
### 2.7 Retrieval diagnostic (optional, RAG only)

Recall@k tells you _whether the right article is in the top-k retrieved chunks_ — independent of whether the LLM then picks the right letter. Without it, you can't tell retrieval failures apart from generation failures.

**Workflow:**

1. **Bootstrap** a labeling stub from your gold set — run the diagnostic cell (bottom of this section). Each row gets the retriever's current top-5 candidate article titles.
2. **Label** — two options:
   - _(Recommended)_ Run **cell 2.7a** to export to Excel. Fill in the highlighted `gold_article_title` column (pick from the 5 candidate columns or type your own title), then run **cell 2.7b** to import back.
   - _(Manual)_ Open `data/eval/retrieval_gold.jsonl` directly and replace each `"gold_article_title": null` with the Wikipedia article title that should be retrieved. Leave `null` to mean "no article suffices" — those rows are skipped in scoring.
   - 50 labels → usable signal; 100+ → comfortable.
3. **Measure**: re-run the diagnostic cell; recall@1/3/5/10 and per-category breakdown come out.

Skip this section entirely if `retriever is None` (you're not running RAG).
"""))

cells.append(code('''
# ── 2.7a  Export retrieval_gold stub → Excel for easy labeling ──────────
# Requires openpyxl: installed below if missing.
import subprocess, sys
subprocess.run([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'], check=True)

import pandas as pd
from openpyxl.styles import PatternFill, Font

RETRIEVAL_GOLD_PATH = PATHS.eval_dir / 'retrieval_gold.jsonl'
EXCEL_PATH          = PATHS.eval_dir / 'retrieval_gold_label.xlsx'

if not RETRIEVAL_GOLD_PATH.exists():
    print('No stub yet — run the diagnostic cell below first to generate retrieval_gold.jsonl.')
else:
    from polimibot.eval.retrieval import load_retrieval_gold
    items = load_retrieval_gold(RETRIEVAL_GOLD_PATH)

    rows = []
    for it in items:
        cands = list(it.candidates) + [''] * 5
        rows.append({
            'question_text':      it.question_text,
            'options':            ' | '.join(it.options),
            'category':           it.category.value if it.category else '',
            'level':              it.level,
            'gold_article_title': it.gold_article_title or '',
            'candidate_1':        cands[0],
            'candidate_2':        cands[1],
            'candidate_3':        cands[2],
            'candidate_4':        cands[3],
            'candidate_5':        cands[4],
        })

    df = pd.DataFrame(rows)
    gold_col_idx = df.columns.get_loc('gold_article_title')  # 0-indexed

    with pd.ExcelWriter(EXCEL_PATH, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Labels')
        ws = writer.sheets['Labels']
        ws.freeze_panes = 'A2'

        yellow = PatternFill(fill_type='solid', fgColor='FFFF00')
        bold   = Font(bold=True)
        ws.cell(row=1, column=gold_col_idx + 1).font = bold
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=gold_col_idx + 1).fill = yellow

        for col_letter, width in [('A', 60), ('B', 40), ('E', 35),
                                   ('F', 30), ('G', 30), ('H', 30), ('I', 30), ('J', 30)]:
            ws.column_dimensions[col_letter].width = width

    n_labeled = sum(1 for it in items if it.is_labeled)
    print(f'Exported {len(items)} rows ({n_labeled} already labeled) → {EXCEL_PATH}')
    print('Fill in the yellow "gold_article_title" column, then run the Import cell below.')
'''))

cells.append(code('''
# ── 2.7b  Import labeled Excel → retrieval_gold.jsonl ───────────────────
import pandas as pd
from dataclasses import replace as dc_replace
from polimibot.eval.retrieval import load_retrieval_gold, save_retrieval_gold

RETRIEVAL_GOLD_PATH = PATHS.eval_dir / 'retrieval_gold.jsonl'
EXCEL_PATH          = PATHS.eval_dir / 'retrieval_gold_label.xlsx'

if not EXCEL_PATH.exists():
    print(f'Excel file not found: {EXCEL_PATH}\\nRun the Export cell above first.')
elif not RETRIEVAL_GOLD_PATH.exists():
    print(f'JSONL stub not found: {RETRIEVAL_GOLD_PATH}\\nRun the diagnostic cell below first.')
else:
    df = pd.read_excel(EXCEL_PATH, sheet_name='Labels', dtype=str).fillna('')
    label_map = {
        row['question_text']: (row['gold_article_title'].strip() or None)
        for _, row in df.iterrows()
    }

    items = load_retrieval_gold(RETRIEVAL_GOLD_PATH)
    n_new = 0
    updated = []
    for it in items:
        new_title = label_map.get(it.question_text, it.gold_article_title)
        if new_title and new_title != it.gold_article_title:
            n_new += 1
        updated.append(dc_replace(it, gold_article_title=new_title))

    save_retrieval_gold(updated, RETRIEVAL_GOLD_PATH)
    n_total = sum(1 for it in updated if it.is_labeled)
    print(f'Imported {n_new} new labels. Total labeled: {n_total} / {len(updated)}')
    print('Re-run the diagnostic cell below to measure Recall@k.')
'''))

cells.append(code('''
# Retrieval diagnostic, optional. Skip if no RAG.
RETRIEVAL_GOLD_PATH = PATHS.eval_dir / 'retrieval_gold.jsonl'
RUN_RETRIEVAL_DIAGNOSTIC = True   # set False to skip

if not RUN_RETRIEVAL_DIAGNOSTIC or retriever is None:
    print('Retrieval diagnostic skipped (RUN_RETRIEVAL_DIAGNOSTIC=False or retriever=None).')
else:
    from polimibot.eval.retrieval import (
        build_labeling_template, save_retrieval_gold,
        load_retrieval_gold, evaluate_retrieval,
        evaluate_retrieval_multi_query,
    )

    if not RETRIEVAL_GOLD_PATH.exists():
        # First time: emit a labeling stub. Open the JSONL and fill in titles.
        stub = build_labeling_template(full, retriever=retriever, k_candidates=5)
        save_retrieval_gold(stub, RETRIEVAL_GOLD_PATH)
        print(f'\\nLabeling stub written → {RETRIEVAL_GOLD_PATH}')
        print('Open it, fill in "gold_article_title" for each row, then re-run this cell.')
    else:
        labeled = load_retrieval_gold(RETRIEVAL_GOLD_PATH)
        n_labeled = sum(1 for it in labeled if it.is_labeled)
        if n_labeled == 0:
            print(f'No labeled rows yet in {RETRIEVAL_GOLD_PATH}.')
            print('Edit the file and fill in "gold_article_title" for each question.')
        else:
            # Dispatch on RAG_USE_MULTI_QUERY so the diagnostic measures the
            # same recipe the runtime RAGStrategy actually uses. The single-
            # query harness was silently mis-measuring when multi-query was on.
            common_name = (
                f'k={RAG_K}'
                + ('+cat' if RAG_USE_CATEGORY_FILTER else '')
                + ('+hybrid' if RAG_USE_HYBRID else '')
                + ('+rerank' if RAG_USE_RERANKER else '')
            )
            if RAG_USE_MULTI_QUERY:
                report = evaluate_retrieval_multi_query(
                    retriever, labeled,
                    ks=(1, 3, 5, 10),
                    use_category_filter=RAG_USE_CATEGORY_FILTER,
                    use_reranker=RAG_USE_RERANKER,
                    use_hybrid=RAG_USE_HYBRID,
                    rerank_oversearch=RERANK_OVERSEARCH,
                    retriever_name=common_name + '+mq',
                )
            else:
                report = evaluate_retrieval(
                    retriever, labeled,
                    ks=(1, 3, 5, 10),
                    use_category_filter=RAG_USE_CATEGORY_FILTER,
                    use_reranker=RAG_USE_RERANKER,
                    use_hybrid=RAG_USE_HYBRID,
                    retriever_name=common_name,
                )
            report.print_summary()
            report.save(PATHS.eval_dir / f'retrieval__{report_id}.json')
'''))

cells.append(obs("Run observations"))

# ────────────────────────── Section 2.8 — Tune RAG hyperparameters ──────
cells.append(md("""
### 2.8 Tune RAG hyperparameters (optional, RAG only)

Sweep a small grid of recipe knobs (`k`, hybrid, reranker, multi-query,
oversearch), measure downstream accuracy on a gold-set slice, then calibrate
one `min_score` threshold **per retrieval path** (`dense` / `rrf` / `rerank`)
from the same sweep — reusing the already-loaded `llm` and `retriever`.

The three threshold knobs are path-aware (one wins per recipe):

| recipe (hybrid, rerank) | path     | knob                        | score scale on `extras.top_score`        |
|-------------------------|----------|-----------------------------|------------------------------------------|
| `(False, False)`        | `dense`  | `RAG_MIN_SCORE`             | cosine ∈ [-1, 1]                         |
| `(True,  False)`        | `rrf`    | `RAG_MIN_SCORE_RRF`         | RRF ∈ ~[0, 0.03]                         |
| `(_,     True)`         | `rerank` | `RAG_MIN_SCORE_RERANK`      | raw cross-encoder logit, model-dependent |

`RAGStrategy` records `extras.top_score` on whichever scale was active, so one
sweep covers all three.

**Workflow**: edit the grid in **cell 2.8a** → run **cell 2.8b** (sweep) →
run **cell 2.8c** (calibrate τ per path) → copy the winner's knob + τ back
into **Section 1.1** and re-run Sections 1.3-1.4 before re-evaluating in 2.2.

Skipped silently when `retriever is None` (no RAG configured) or `USE_MOCK=True`.
"""))

cells.append(md("#### 2.8a Sweep grid"))

cells.append(code('''
# Keep the grid small to start; expand once you've seen the leaderboard.
# Each combo runs evaluate_strategy on a gold slice — wall time scales as
# (n_combos × SWEEP_N × per-question latency). For a quick first pass:
# SWEEP_N=30, default grid → ~30 valid combos, ~10–20 minutes on Colab T4.
SWEEP_KS                = [3, 5, 8]
SWEEP_HYBRID            = [False, True]      # dense vs hybrid RRF
SWEEP_RERANK            = [False, True]      # off vs cross-encoder
SWEEP_MULTI_QUERY       = [False, True]      # 1 query vs 1+4 queries
SWEEP_RERANK_OVERSEARCH = [3, 5]             # only varied when rerank=True
SWEEP_N                 = 50                  # gold-set slice size
'''))

cells.append(md("#### 2.8b Run the sweep"))

cells.append(code('''
import itertools
import time as _time
import pandas as pd
from polimibot.eval.evaluator import evaluate_strategy
from polimibot.strategies.rag_strategy import RAGStrategy

if retriever is None or USE_MOCK:
    print('Tuning sweep skipped (retriever is None or USE_MOCK=True).')
    sweep_df = None
    sweep_samples_by_combo = {}
else:
    # Deterministic slice — same prefix every time so reruns are comparable.
    _gold_slice = list(gold)[:SWEEP_N]
    sweep_samples_by_combo: dict[tuple, list] = {}

    rows = []
    combos = list(itertools.product(
        SWEEP_KS, SWEEP_HYBRID, SWEEP_RERANK,
        SWEEP_MULTI_QUERY, SWEEP_RERANK_OVERSEARCH,
    ))
    n_valid = 0
    n_total = 0
    for k_, hyb, rer, mq, os_ in combos:
        n_total += 1
        # Skip combos that can't compose under the current retriever.
        if rer and not retriever.has_reranker:
            continue
        if hyb and not retriever.has_bm25:
            continue
        # rerank_oversearch is a no-op when rerank=False — pick one canonical
        # value so we don't double-count identical recipes.
        if not rer and os_ != SWEEP_RERANK_OVERSEARCH[0]:
            continue
        n_valid += 1

        # Reuse the shared rag kwargs from Section 1.4; override the swept
        # knobs and turn thresholds OFF (we want the full score distribution
        # to feed 2.8c calibration).
        _kw = dict(_rag_kwargs)
        _kw.update(
            k=k_, use_hybrid=hyb, use_reranker=rer,
            use_multi_query=mq, rerank_oversearch=os_,
            min_score=None, min_score_rrf=None, min_score_rerank=None,
            # Disable live fallback during the sweep so timing is offline-only
            # and score distributions reflect the local index, not Wikipedia.
            use_live_fallback=False,
            index_grower=None,
        )
        strat = RAGStrategy(llm, retriever, **_kw)
        t0 = _time.monotonic()
        rep = evaluate_strategy(strat, _gold_slice, verbose=False)
        dt = _time.monotonic() - t0

        # Stash the per-question samples for 2.8c — calibrating from in-memory
        # samples avoids the JSONL roundtrip.
        sweep_samples_by_combo[(k_, hyb, rer, mq, os_)] = rep.samples

        rows.append({
            'k': k_, 'hybrid': hyb, 'rerank': rer, 'mq': mq, 'os': os_,
            'accuracy': round(rep.accuracy, 4),
            'ece':      round(rep.ece, 4),
            'wall_s':   round(dt, 1),
            'n':        rep.n_total,
        })
        # Live progress so the user can ctrl-c if a combo is too slow.
        print(f'  k={k_} hyb={int(hyb)} rer={int(rer)} mq={int(mq)} os={os_}'
              f'  acc={rep.accuracy:.1%}  ({dt:.1f}s)')

    sweep_df = pd.DataFrame(rows).sort_values('accuracy', ascending=False)
    print(f'\\nRan {n_valid} / {n_total} combos (others skipped: no reranker/bm25).')
    print('\\nTop 10 RAG recipes by accuracy on the gold slice:')
    print(sweep_df.head(10).to_string(index=False))
    if not sweep_df.empty:
        win = sweep_df.iloc[0]
        print(f'\\nWinner: k={int(win.k)}, hybrid={bool(win.hybrid)}, '
              f'rerank={bool(win.rerank)}, mq={bool(win.mq)}, os={int(win.os)}  '
              f'→ acc={win.accuracy:.1%}')
'''))

cells.append(md("""#### 2.8c Calibrate `RAG_MIN_SCORE`, `RAG_MIN_SCORE_RRF`, `RAG_MIN_SCORE_RERANK`

For each retrieval path that ran in the sweep, pick the best-accuracy recipe
in that path, then maximise gated-policy expected accuracy:

```
expected_acc(τ) = acc(score ≥ τ) · P(score ≥ τ)
                + bare_baseline_acc · P(score < τ)
```

The bare-baseline accuracy comes from running the LLM **without RAG** on the
same gold slice — measured automatically in this cell.

The cell outputs one row per path with a calibrated τ ready to paste back into
Section 1.1. You only need to set the knob for the recipe you'll ship."""))

cells.append(code('''
from polimibot.eval.threshold_calibration import calibrate_threshold
from polimibot.strategies.llm_baseline import BaselineLLMStrategy

if sweep_df is None or sweep_df.empty:
    print('No sweep results — run cell 2.8b first.')
    rec_df = None
else:
    # 1. Bare-baseline accuracy on the same slice (no RAG).
    #    Drives the calibration formula's fallback-on-gated assumption.
    _baseline_for_cal = BaselineLLMStrategy(
        llm, style=PROMPT_STYLE,
        use_score_options=USE_SCORE_OPTIONS,
        direct_max_new_tokens=DIRECT_MAX_NEW_TOKENS,
        cot_max_new_tokens=COT_MAX_NEW_TOKENS,
        stop_strings=STOP_STRINGS,
    )
    _bare_rep = evaluate_strategy(_baseline_for_cal, list(gold)[:SWEEP_N], verbose=False)
    BARE_BASELINE_ACC = _bare_rep.accuracy
    print(f'Bare-baseline accuracy on the slice: {BARE_BASELINE_ACC:.2%}')

    # 2. Bucket sweep rows by retrieval path. Within each path, pick the
    #    best-accuracy recipe and calibrate τ from its in-memory samples.
    def _path_of(row):
        if row.rerank: return 'rerank'
        if row.hybrid: return 'rrf'
        return 'dense'
    sweep_df['path'] = sweep_df.apply(_path_of, axis=1)

    _knob_for_path = {
        'dense':  'RAG_MIN_SCORE',
        'rrf':    'RAG_MIN_SCORE_RRF',
        'rerank': 'RAG_MIN_SCORE_RERANK',
    }
    recs = []
    for path, group in sweep_df.groupby('path'):
        best = group.sort_values('accuracy', ascending=False).iloc[0]
        key = (int(best.k), bool(best.hybrid), bool(best.rerank),
               bool(best.mq), int(best['os']))
        samples = sweep_samples_by_combo.get(key, [])
        pairs = [(s.extras.get('top_score'), s.correct) for s in samples
                 if s.extras.get('top_score') is not None]
        if not pairs:
            continue   # nothing to calibrate on this path
        scores  = [s for s, _ in pairs]
        corrects = [c for _, c in pairs]
        cand = calibrate_threshold(
            scores, corrects, bare_baseline_acc=BARE_BASELINE_ACC,
        )
        top = cand[0]
        recs.append({
            'path':                path,
            'knob':                _knob_for_path[path],
            'best_recipe':         f'k={int(best.k)} mq={int(best.mq)} os={int(best["os"])}',
            'recipe_accuracy':     round(best.accuracy, 4),
            'tau':                 round(top['tau'], 4),
            'expected_acc_at_tau': round(top['expected'], 4),
            'p_ungated_at_tau':    round(top['p_ungated'], 3),
            'n_pairs':             len(pairs),
        })

    rec_df = pd.DataFrame(recs).sort_values('expected_acc_at_tau', ascending=False)
    print('\\nCalibrated thresholds (one per retrieval path):')
    print(rec_df.to_string(index=False))
    print('\\nPaste the τ matching your shipped recipe into Section 1.1, e.g.:')
    for r in recs:
        print(f"  {r['knob']} = {r['tau']}")
    print()
    print('Note: this assumes the bare-LLM baseline catches gated questions.')
    print('If you set USE_LIVE_FALLBACK=True, the live arm catches them instead —')
    print('so the calibrated τ is a lower bound; raising it slightly is usually')
    print('fine and reduces Wikipedia API load.')
'''))

cells.append(obs("Tuning observations"))

# ────────────────────────── Section 3 — Compare ──────────────────────────
cells.append(md("""
---

## 3. Compare

Read every saved report from `data/eval/*.json` into one comparison table. This is your leaderboard — it grows automatically as you save more reports in Section 2.
"""))

cells.append(md("### 3.1 Build the leaderboard"))

cells.append(code('''
# Existing build_leaderboard reads every *.json in data/eval/, this it does.
from polimibot.eval.make_leaderboard import build_leaderboard

leaderboard = build_leaderboard(PATHS.eval_dir)
if leaderboard.empty:
    print('No reports yet. Run Section 2 with a strategy first.')
else:
    print(f'\\nLeaderboard ({len(leaderboard)} strategies):')
leaderboard
'''))

cells.append(md("### 3.2 Save the leaderboard"))

cells.append(code('''
# Persistent CSV, reproducible the table is.
if not leaderboard.empty:
    save_results(leaderboard, 'leaderboard')
'''))

cells.append(md("### 3.3 Comparison plot"))

cells.append(code('''
# Strategies on the x-axis; accuracy bars sorted descending. Speak louder than tables, plots do.
if leaderboard.empty:
    print('Leaderboard empty — nothing to plot yet.')
else:
    fig, (ax_acc, ax_lat) = plt.subplots(1, 2, figsize=(13, 4))

    df_sorted = leaderboard.sort_values('accuracy', ascending=False)
    ax_acc.barh(df_sorted['strategy'], df_sorted['accuracy'], color='steelblue', edgecolor='white')
    ax_acc.set_xlim(0, 1.0)
    ax_acc.set_xlabel('Accuracy')
    ax_acc.invert_yaxis()
    ax_acc.axvline(0.25, linestyle='--', color='grey', linewidth=1, label='random')
    ax_acc.set_title('Accuracy by strategy')
    ax_acc.legend(loc='lower right')

    ax_lat.barh(df_sorted['strategy'], df_sorted['latency_p50_s'], color='salmon', edgecolor='white',
                label='p50')
    ax_lat.barh(df_sorted['strategy'], df_sorted['latency_p95_s'] - df_sorted['latency_p50_s'],
                left=df_sorted['latency_p50_s'], color='lightsalmon', edgecolor='white',
                label='p50→p95')
    ax_lat.invert_yaxis()
    ax_lat.set_xlabel('Seconds per question')
    ax_lat.set_title('Latency (p50 + p95 tail)')
    ax_lat.legend(loc='lower right')

    plt.tight_layout()
    plt.show()
'''))

cells.append(md("### 3.4 Per-category cross-strategy heatmap"))

cells.append(code('''
# Reports include per-category accuracy. Heatmap, build it from disk we do.
def _load_report(p: Path) -> dict:
    return json.loads(p.read_text())

rows = []
for p in sorted(PATHS.eval_dir.glob('*.json')):
    data = _load_report(p)
    for cat, stats in data.get('by_category', {}).items():
        rows.append({
            'strategy': data['strategy_name'],
            'category': cat,
            'accuracy': stats.get('accuracy', float('nan')),
            'n':        stats.get('n', 0),
        })

if rows:
    heat = (pd.DataFrame(rows)
              .pivot_table(index='strategy', columns='category', values='accuracy'))
    fig, ax = plt.subplots(figsize=(7, max(2, 0.5 * len(heat))))
    im = ax.imshow(heat.values, vmin=0, vmax=1, cmap='Blues', aspect='auto')
    ax.set_xticks(range(len(heat.columns))); ax.set_xticklabels(heat.columns, rotation=30, ha='right')
    ax.set_yticks(range(len(heat.index)));   ax.set_yticklabels(heat.index)
    for i in range(heat.shape[0]):
        for j in range(heat.shape[1]):
            v = heat.values[i, j]
            if pd.notna(v):
                ax.text(j, i, f'{v:.0%}', ha='center', va='center',
                        color='white' if v > 0.55 else 'black', fontsize=9)
    fig.colorbar(im, ax=ax, fraction=0.04, pad=0.02)
    ax.set_title('Per-category accuracy across strategies')
    plt.tight_layout()
    plt.show()
    save_results(heat.reset_index(), 'percategory_matrix')
else:
    print('No per-category data yet.')
'''))

cells.append(obs("Comparison observations"))

# ────────────────────────── Section 4 — Save ──────────────────────────
cells.append(md("""
---

## 4. Save

Consolidated artefacts for the write-up. All previous sections write incrementally; this section just summarises what's on disk.
"""))

cells.append(md("### 4.1 Inventory"))

cells.append(code('''
# What did we produce? List, count, recap.
def _list(dir_: Path, pattern: str) -> list[Path]:
    return sorted(dir_.glob(pattern)) if dir_.exists() else []

run_logs    = _list(PATHS.runs_dir,           '*.jsonl')
eval_jsons  = _list(PATHS.eval_dir,           '*.json')
result_csvs = _list(RESULTS_DIR,              '*.csv') if RESULTS_DIR.exists() else []
plots       = _list(PATHS.eval_dir,           '*.png')

inventory = pd.DataFrame({
    'kind':  ['run logs (JSONL)', 'eval reports (JSON)', 'results (CSV)', 'plots (PNG)'],
    'count': [len(run_logs), len(eval_jsons), len(result_csvs), len(plots)],
})
print(inventory.to_string(index=False))
'''))

cells.append(md("### 4.2 Final summary"))

cells.append(code('''
# A concise text summary, useful to paste into the report it is.
print('═' * 60)
print(f'PoliMillionaire — experiment summary')
print('═' * 60)
print(f'API URL          : {RUNTIME.api_url}')
print(f'Model            : {MODEL_ID}  (mock={USE_MOCK})')
print(f'Strategy         : {strategy.name}')
if not leaderboard.empty:
    top = leaderboard.iloc[0]
    print(f"Leader strategy  : {top['strategy']}")
    print(f"Leader accuracy  : {top['accuracy']:.1%}")
    print(f"Leader p50/p95   : {top['latency_p50_s']:.2f}s / {top['latency_p95_s']:.2f}s")
print(f'Reports on disk  : {len(eval_jsons)}')
print('═' * 60)
'''))

cells.append(obs("Save / final observations"))

# ────────────────────────── VRAM hygiene helper section ──────────────────────────
cells.append(md("""
---

## Appendix — VRAM hygiene

When you finish a model and want to load a different one, run the cell below first. T4 GPUs in Colab die without warning if VRAM gets fragmented.
"""))

cells.append(code('''
# Free the current LLM completely. The next Section 1.2 run will reload from scratch.
if 'llm' in globals():
    unload_llm(llm)
    del llm
    if 'LOADED_MODEL_ID' in globals():
        del LOADED_MODEL_ID
clear_vram()
print('VRAM cleared. Edit MODEL_ID in Section 1.1 and re-run Sections 1.2 → 2.')
'''))


# ──────────────────────────────────────────────────────────────────────
# Assemble + write
# ──────────────────────────────────────────────────────────────────────

nb = nbf.v4.new_notebook()
nb.cells = cells
nb.metadata.update({
    'kernelspec': {
        'display_name': 'Python 3',
        'language':     'python',
        'name':         'python3',
    },
    'language_info': {
        'name':    'python',
        'version': '3.11',
    },
    'colab': {
        'name':       'PoliMillionaire.ipynb',
        'provenance': [],
    },
})

OUT = Path(__file__).resolve().parent.parent / 'PoliMillionaire.ipynb'
nbf.write(nb, OUT)

# Sanity: read back, check every code cell is valid Python (after stripping
# IPython magics — lines starting with % or ! aren't Python and ast.parse
# would reject them, but Jupyter handles them fine at runtime).
import ast
import re

def _strip_ipython_magics(src: str) -> str:
    return re.sub(r'^[ \t]*[%!].*$', '', src, flags=re.MULTILINE)

nb_check = nbf.read(OUT, as_version=4)
n_md = sum(1 for c in nb_check.cells if c.cell_type == 'markdown')
n_code = sum(1 for c in nb_check.cells if c.cell_type == 'code')
errors: list[str] = []
for i, c in enumerate(nb_check.cells):
    if c.cell_type == 'code':
        try:
            ast.parse(_strip_ipython_magics(c.source))
        except SyntaxError as e:
            errors.append(f'cell {i}: {e}')

print(f'\\nWrote {OUT}')
print(f'  cells: {len(nb.cells)} ({n_code} code, {n_md} markdown)')
if errors:
    print('  ✗ syntax errors:')
    for e in errors:
        print(f'    {e}')
    raise SystemExit(1)
print('  OK all code cells parse')
